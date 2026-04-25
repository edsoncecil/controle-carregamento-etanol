import io
from datetime import datetime

from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from urllib.parse import urlencode

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib.styles import getSampleStyleSheet

from .forms import CarregamentoForm
from .models import Carregamento


def logout_view(request):
    logout(request)
    return render(request, "registration/logged_out.html")


STATUS_FILTER_OPTIONS = [
    ("ATIVOS", "Somente ativos (esconde finalizados)"),
    ("TODOS", "Todos os status"),
    ("CARREGANDO", "Apenas carregando"),
    ("PATIO", "Apenas pátio"),
    ("FINALIZADO", "Apenas finalizado"),
]
PER_PAGE_OPTIONS = [10, 20, 50, 100]

SORT_FIELDS = {
    "motorista": "motorista",
    "placa": "placa",
    "distribuidora": "distribuidora",
    "transportador": "transportador",
    "litragem": "litragem_sem_excesso",
    "ordem": "ordem",
    "situacao": "situacao",
    "lacres": "lacres",
    "cadastro": "data_hora_cadastro",
}


def _sanitize_sort_by(raw_sort_by):
    if raw_sort_by in SORT_FIELDS:
        return raw_sort_by
    return "cadastro"


def _sanitize_sort_dir(raw_sort_dir):
    if raw_sort_dir == "asc":
        return "asc"
    return "desc"


def _sanitize_per_page(raw_per_page):
    try:
        value = int(raw_per_page)
    except (TypeError, ValueError):
        return 20
    if value in PER_PAGE_OPTIONS:
        return value
    return 20


def _build_query_string(status_filtro, busca, sort_by, sort_dir, per_page, page=None):
    query = {
        "status": status_filtro,
        "q": busca,
        "sort": sort_by,
        "dir": sort_dir,
        "per_page": per_page,
    }
    if page is not None:
        query["page"] = page
    return urlencode(query)


def _build_sort_links(status_filtro, busca, sort_by, sort_dir, per_page):
    links = {}
    for key in SORT_FIELDS:
        next_dir = "asc"
        if sort_by == key and sort_dir == "asc":
            next_dir = "desc"
        links[key] = _build_query_string(status_filtro, busca, key, next_dir, per_page)
    return links


def _listar_carregamentos(status_filtro, busca, sort_by, sort_dir):
    carregamentos = Carregamento.objects.all()

    if status_filtro == "ATIVOS":
        carregamentos = carregamentos.exclude(situacao="FINALIZADO")
    elif status_filtro != "TODOS":
        carregamentos = carregamentos.filter(situacao=status_filtro)

    busca = busca.strip()
    if busca:
        filtro_busca = (
            Q(motorista__icontains=busca)
            | Q(placa__icontains=busca)
            | Q(distribuidora__icontains=busca)
            | Q(transportador__icontains=busca)
            | Q(ordem__icontains=busca)
            | Q(situacao__icontains=busca)
        )
        if busca.isdigit():
            valor_numerico = int(busca)
            filtro_busca |= Q(lacres=valor_numerico) | Q(litragem_sem_excesso=valor_numerico)
        carregamentos = carregamentos.filter(filtro_busca)

    sort_field = SORT_FIELDS[sort_by]
    order_by = sort_field if sort_dir == "asc" else f"-{sort_field}"
    carregamentos = carregamentos.order_by(order_by, "-id")

    return carregamentos


def _sanitize_status_filter(raw_status):
    allowed_filters = {opt[0] for opt in STATUS_FILTER_OPTIONS}
    if raw_status in allowed_filters:
        return raw_status
    return "ATIVOS"


def _render_fila(
    request,
    form,
    carregamento_em_edicao=None,
    status_filtro="ATIVOS",
    busca="",
    sort_by="cadastro",
    sort_dir="asc",
    per_page=20,
    page=1,
):
    carregamentos_qs = _listar_carregamentos(status_filtro, busca, sort_by, sort_dir)
    paginator = Paginator(carregamentos_qs, per_page)
    page_obj = paginator.get_page(page)
    carregamentos = page_obj.object_list

    query_string = _build_query_string(status_filtro, busca, sort_by, sort_dir, per_page, page_obj.number)
    base_query_string = _build_query_string(status_filtro, busca, sort_by, sort_dir, per_page)
    sort_links = _build_sort_links(status_filtro, busca, sort_by, sort_dir, per_page)

    return render(
        request,
        "core/fila_carregamento.html",
        {
            "form": form,
            "carregamentos": carregamentos,
            "page_obj": page_obj,
            "carregamento_em_edicao": carregamento_em_edicao,
            "status_filtro": status_filtro,
            "status_filter_options": STATUS_FILTER_OPTIONS,
            "per_page_options": PER_PAGE_OPTIONS,
            "busca": busca,
            "sort_by": sort_by,
            "sort_dir": sort_dir,
            "per_page": per_page,
            "sort_links": sort_links,
            "query_string": query_string,
            "base_query_string": base_query_string,
        },
    )


@login_required
def fila_carregamento(request):
    status_filtro = _sanitize_status_filter(request.GET.get("status", "ATIVOS"))
    busca = request.GET.get("q", "").strip()
    sort_by = _sanitize_sort_by(request.GET.get("sort", "cadastro"))
    sort_dir = _sanitize_sort_dir(request.GET.get("dir", "asc"))
    per_page = _sanitize_per_page(request.GET.get("per_page", "20"))
    page = request.GET.get("page", "1")

    if request.method == "POST":
        status_filtro = _sanitize_status_filter(request.POST.get("status_filtro", "ATIVOS"))
        busca = request.POST.get("busca", "").strip()
        sort_by = _sanitize_sort_by(request.POST.get("sort_by", "cadastro"))
        sort_dir = _sanitize_sort_dir(request.POST.get("sort_dir", "asc"))
        per_page = _sanitize_per_page(request.POST.get("per_page", "20"))
        page = request.POST.get("page", "1")
        form = CarregamentoForm(request.POST)
        if form.is_valid():
            carregamento = form.save(commit=False)
            carregamento.alterado_por = str(request.user)
            carregamento.data_hora_alteracao = timezone.now()
            carregamento.save()
            messages.success(
                request,
                f"Carregamento de {carregamento.motorista} ({carregamento.placa}) cadastrado com sucesso.",
            )
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True, "message": f"Carregamento de {carregamento.motorista} ({carregamento.placa}) cadastrado com sucesso."})
            query_string = _build_query_string(status_filtro, busca, sort_by, sort_dir, per_page, page)
            return redirect(f"{reverse('fila_carregamento')}?{query_string}")
    else:
        form = CarregamentoForm()

    return _render_fila(
        request,
        form,
        status_filtro=status_filtro,
        busca=busca,
        sort_by=sort_by,
        sort_dir=sort_dir,
        per_page=per_page,
        page=page,
    )


@login_required
def editar_carregamento(request, pk):
    carregamento = get_object_or_404(Carregamento, pk=pk)
    status_filtro = _sanitize_status_filter(request.GET.get("status", "ATIVOS"))
    busca = request.GET.get("q", "").strip()
    sort_by = _sanitize_sort_by(request.GET.get("sort", "cadastro"))
    sort_dir = _sanitize_sort_dir(request.GET.get("dir", "asc"))
    per_page = _sanitize_per_page(request.GET.get("per_page", "20"))
    page = request.GET.get("page", "1")

    if request.method == "POST":
        status_filtro = _sanitize_status_filter(request.POST.get("status_filtro", "ATIVOS"))
        busca = request.POST.get("busca", "").strip()
        sort_by = _sanitize_sort_by(request.POST.get("sort_by", "cadastro"))
        sort_dir = _sanitize_sort_dir(request.POST.get("sort_dir", "asc"))
        per_page = _sanitize_per_page(request.POST.get("per_page", "20"))
        page = request.POST.get("page", "1")
        form = CarregamentoForm(request.POST, instance=carregamento)
        if form.is_valid():
            carregamento = form.save(commit=False)
            carregamento.alterado_por = str(request.user)
            carregamento.data_hora_alteracao = timezone.now()
            carregamento.save()
            messages.success(
                request,
                f"Registro de {carregamento.motorista} ({carregamento.placa}) atualizado com sucesso.",
            )
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True, "message": f"Registro de {carregamento.motorista} ({carregamento.placa}) atualizado com sucesso."})
            query_string = _build_query_string(status_filtro, busca, sort_by, sort_dir, per_page, page)
            return redirect(f"{reverse('fila_carregamento')}?{query_string}")
    else:
        form = CarregamentoForm(instance=carregamento)

    return _render_fila(
        request,
        form,
        carregamento_em_edicao=carregamento,
        status_filtro=status_filtro,
        busca=busca,
        sort_by=sort_by,
        sort_dir=sort_dir,
        per_page=per_page,
        page=page,
    )


@login_required
def api_carregamentos(request):
    status_filtro = _sanitize_status_filter(request.GET.get("status", "ATIVOS"))
    busca = request.GET.get("q", "").strip()
    sort_by = _sanitize_sort_by(request.GET.get("sort", "cadastro"))
    sort_dir = _sanitize_sort_dir(request.GET.get("dir", "asc"))
    per_page = _sanitize_per_page(request.GET.get("per_page", "20"))
    page = request.GET.get("page", "1")

    carregamentos_qs = _listar_carregamentos(status_filtro, busca, sort_by, sort_dir)
    paginator = Paginator(carregamentos_qs, per_page)
    page_obj = paginator.get_page(page)

    query_string = _build_query_string(status_filtro, busca, sort_by, sort_dir, per_page)
    sort_links = _build_sort_links(status_filtro, busca, sort_by, sort_dir, per_page)

    html = render_to_string(
        "core/_tabela_carregamentos.html",
        {
            "carregamentos": page_obj.object_list,
            "page_obj": page_obj,
            "status_filtro": status_filtro,
            "status_filter_options": STATUS_FILTER_OPTIONS,
            "per_page_options": PER_PAGE_OPTIONS,
            "busca": busca,
            "sort_by": sort_by,
            "sort_dir": sort_dir,
            "per_page": per_page,
            "sort_links": sort_links,
            "query_string": query_string,
            "base_query_string": _build_query_string(status_filtro, busca, sort_by, sort_dir, per_page),
        },
    )

    return JsonResponse({"html": html, "success": True})


# ---------------------------------------------------------------------------
# Helpers para relatórios
# ---------------------------------------------------------------------------

RELATORIO_COLUNAS = [
    "#",
    "Motorista",
    "Placa",
    "Distribuidora",
    "Transportador",
    "Litragem s/ Excesso",
    "Ordem",
    "Situação",
    "Lacres",
    "Data/Hora Cadastro",
    "Última Alteração",
]


def _parse_datas(request):
    """Retorna (data_inicio, data_fim, qs_filtrado) ou levanta ValueError."""
    data_inicio_str = request.GET.get("data_inicio", "").strip()
    data_fim_str = request.GET.get("data_fim", "").strip()

    if not data_inicio_str or not data_fim_str:
        raise ValueError("Informe a data de início e a data de fim.")

    data_inicio = datetime.strptime(data_inicio_str, "%Y-%m-%d")
    data_fim = datetime.strptime(data_fim_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59)

    tz = timezone.get_current_timezone()
    data_inicio = timezone.make_aware(data_inicio, tz)
    data_fim = timezone.make_aware(data_fim, tz)

    qs = Carregamento.objects.filter(
        data_hora_cadastro__gte=data_inicio,
        data_hora_cadastro__lte=data_fim,
    ).order_by("data_hora_cadastro", "id")

    return data_inicio, data_fim, qs


def _carregamento_to_row(idx, c):
    """Converte um Carregamento em lista de valores para relatório."""
    ultima = "-"
    if c.alterado_por and c.data_hora_alteracao:
        ultima = f"{c.alterado_por.upper()} - {timezone.localtime(c.data_hora_alteracao).strftime('%d/%m/%Y %H:%M')}"

    return [
        idx,
        c.motorista.upper(),
        c.placa.upper(),
        c.distribuidora.upper(),
        c.transportador.upper(),
        c.litragem_sem_excesso,
        c.ordem,
        c.get_situacao_display(),
        c.lacres,
        timezone.localtime(c.data_hora_cadastro).strftime("%d/%m/%Y %H:%M"),
        ultima,
    ]


# ---------------------------------------------------------------------------
# View: Relatório XLSX
# ---------------------------------------------------------------------------

@login_required
def relatorio_xlsx(request):
    try:
        data_inicio, data_fim, qs = _parse_datas(request)
    except ValueError as e:
        return HttpResponse(str(e), status=400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"

    # --- Título mesclado ---
    periodo = f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(RELATORIO_COLUNAS))
    titulo_cell = ws.cell(row=1, column=1, value=f"Relatório de Carregamentos — {periodo}")
    titulo_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    titulo_cell.fill = PatternFill(start_color="0E7490", end_color="0E7490", fill_type="solid")
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # --- Cabeçalho ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="124E63", end_color="124E63", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D4DCE6"),
        right=Side(style="thin", color="D4DCE6"),
        top=Side(style="thin", color="D4DCE6"),
        bottom=Side(style="thin", color="D4DCE6"),
    )

    for col_idx, col_name in enumerate(RELATORIO_COLUNAS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- Dados ---
    row_fill_even = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    data_alignment = Alignment(horizontal="center", vertical="center")

    for idx, c in enumerate(qs, start=1):
        row_data = _carregamento_to_row(idx, c)
        row_num = idx + 2  # +2 porque linha 1=título, linha 2=cabeçalho
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border
            if idx % 2 == 0:
                cell.fill = row_fill_even

    # --- Se não há dados ---
    if not qs.exists():
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(RELATORIO_COLUNAS))
        empty_cell = ws.cell(row=3, column=1, value="Nenhum carregamento encontrado no período selecionado.")
        empty_cell.alignment = Alignment(horizontal="center")
        empty_cell.font = Font(italic=True, color="666666")

    # --- Ajustar largura das colunas ---
    col_widths = [5, 30, 12, 20, 20, 18, 10, 14, 8, 18, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # --- Gerar resposta ---
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"relatorio_carregamentos_{data_inicio.strftime('%d%m%Y')}_{data_fim.strftime('%d%m%Y')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# View: Relatório PDF
# ---------------------------------------------------------------------------

@login_required
def relatorio_pdf(request):
    try:
        data_inicio, data_fim, qs = _parse_datas(request)
    except ValueError as e:
        return HttpResponse(str(e), status=400)

    buffer = io.BytesIO()
    periodo = f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title="Relatório de Carregamentos",
    )

    styles = getSampleStyleSheet()
    elements = []

    # --- Título ---
    title_style = styles["Title"].clone("titulo_relatorio")
    title_style.fontSize = 16
    title_style.textColor = colors.HexColor("#073B4C")
    title_style.spaceAfter = 4
    elements.append(Paragraph("Relatório de Carregamentos de Etanol", title_style))

    subtitle_style = styles["Normal"].clone("subtitulo_relatorio")
    subtitle_style.fontSize = 11
    subtitle_style.textColor = colors.HexColor("#334155")
    subtitle_style.alignment = 1  # center
    subtitle_style.spaceAfter = 14
    elements.append(Paragraph(f"Período: {periodo}", subtitle_style))

    # --- Montar dados da tabela ---
    table_data = [RELATORIO_COLUNAS]

    if qs.exists():
        for idx, c in enumerate(qs, start=1):
            table_data.append(_carregamento_to_row(idx, c))
    else:
        table_data.append(["—"] * len(RELATORIO_COLUNAS))

    # --- Larguras proporcionais (landscape A4 ~= 277mm útil) ---
    available_width = landscape(A4)[0] - 24 * mm
    col_ratios = [0.03, 0.14, 0.06, 0.10, 0.10, 0.09, 0.05, 0.07, 0.04, 0.10, 0.22]
    col_widths = [available_width * r for r in col_ratios]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # --- Estilos da tabela ---
    header_bg = colors.HexColor("#124E63")
    alt_row_bg = colors.HexColor("#F2F5F8")
    line_color = colors.HexColor("#D4DCE6")

    style_commands = [
        # Cabeçalho
        ("BACKGROUND", (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Corpo
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 1), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),  # motorista à esquerda
        # Grid
        ("GRID", (0, 0), (-1, -1), 0.5, line_color),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, alt_row_bg]),
        # Padding
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]

    table.setStyle(TableStyle(style_commands))
    elements.append(table)

    # --- Rodapé informativo ---
    elements.append(Spacer(1, 10 * mm))
    footer_style = styles["Normal"].clone("rodape_relatorio")
    footer_style.fontSize = 8
    footer_style.textColor = colors.HexColor("#64748B")
    gerado_em = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")
    total = qs.count()
    elements.append(
        Paragraph(f"Gerado em: {gerado_em} &nbsp;|&nbsp; Total de registros: {total}", footer_style)
    )

    doc.build(elements)
    buffer.seek(0)

    filename = f"relatorio_carregamentos_{data_inicio.strftime('%d%m%Y')}_{data_fim.strftime('%d%m%Y')}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
